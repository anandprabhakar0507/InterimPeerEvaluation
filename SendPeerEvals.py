# -*- coding: utf-8 -*-
"""
Created on Wed Sep 27 10:34:16 2017

@author: rmeuth
"""

import win32com.client as win32
import openpyxl





class Student_Review:
    def __init__(self, init_name):
        self.name = init_name
        self.contact = "" # email address
        self.attendance = 0
        self.discussions = 0
        self.timeliness = 0
        self.quality = 0
        self.supportive = 0
        self.contribution = 0
        self.comments = []
        self.review_count = 0
    
    def add_review(self, attendance, 
                   discussions, 
                   timeliness, 
                   quality, 
                   supportive, 
                   contribution, 
                   comment):
        self.attendance += attendance
        self.timeliness += timeliness
        self.discussions += discussions
        self.quality += quality
        self.supportive += supportive
        self.contribution += contribution
        if comment is not None:
            self.comments.append(comment)
        self.review_count += 1
        
    def add_email(self, username):
        self.contact = username + "@asu.edu"
    
    def finalize(self):
        self.attendance /= self.review_count
        self.discussions /= self.review_count
        self.timeliness /= self.review_count
        self.quality /= self.review_count
        self.supportive /= self.review_count
        self.contribution /= self.review_count
        
        

class Eval_Reader:
    def __init__(self, filename):
        self.wb = openpyxl.load_workbook(filename)
        self.eval_sheet = self.wb.get_sheet_by_name("Form Responses 1")
        self.contact_sheet = self.wb.get_sheet_by_name("Contacts")
        self.reviews = {} # dictionary of name-review pairs.
        self.outlook = win32.Dispatch('outlook.application')
         
    def update_contacts(self):
        current_row = 2 # 1-based, skipping header row
        while current_row <= self.contact_sheet.max_row:
            name = self.contact_sheet.cell(row=current_row, column=1).value + " " + self.contact_sheet.cell(row=current_row, column=2).value
            username = self.contact_sheet.cell(row=current_row, column=3).value
            #print(name + " - " + username + "@asu.edu")
            if not name in self.reviews:
                self.reviews[name] = Student_Review(name)
            
            self.reviews[name].add_email(username)
            current_row += 1
        
    def update_reviews(self):
        current_row = 2 # 1-based, skipping header row
        while current_row <= self.eval_sheet.max_row:
            name = self.eval_sheet.cell(row=current_row, column=3).value 
            if not name in self.reviews:
                self.reviews[name] = Student_Review(name)
                #print(name)
            
            attendance = self.eval_sheet.cell(row=current_row, column=4).value
            discussions = self.eval_sheet.cell(row=current_row, column=5).value
            timeliness = self.eval_sheet.cell(row=current_row, column=6).value
            quality = self.eval_sheet.cell(row=current_row, column=7).value
            supportive = self.eval_sheet.cell(row=current_row, column=8).value
            contribution = self.eval_sheet.cell(row=current_row, column=9).value
            comment = self.eval_sheet.cell(row=current_row, column=10).value
            
            self.reviews[name].add_review(attendance, 
                   discussions, 
                   timeliness, 
                   quality, 
                   supportive, 
                   contribution, 
                   comment)
            current_row += 1
            
    def finalize_reviews(self):
        for name in self.reviews:
            self.reviews[name].finalize()
    
    def format_review(self, review):
        message = ""
        message += "Review summary for " + review.name + ":\n"
        message += "Review Count: " + str(review.review_count) + "\n"
        message += "Average Scores:\n"
        message += "\tAttendance:\t" + str(review.attendance) + "\n"
        message += "\tDiscussions:\t" + str(review.discussions) + "\n"
        message += "\tTimeliness:\t" + str(review.timeliness) + "\n"
        message += "\tWork Quality:\t" + str(review.quality) + "\n"
        message += "\tSupportive:\t" + str(review.supportive) + "\n"
        message += "\tContribution:\t" + str(review.contribution) + "\n"
        message += "\nComments:\n"
        for comment in review.comments:
            message += "\t" + comment + "\n"
        return message
            
    def write_email_file(self, filename):
        file = open(filename, 'w')
        for name in self.reviews:
            file.write("REVIEWEE:\t" + name + "\n")
            file.write("DESTINATION:\t" + self.reviews[name].contact + "\n")
            file.write("SUBJECT:\tFSE100 Peer Review Summary\n" )
            file.write(self.format_review(self.reviews[name]))
            file.write("--"*20 + "\n\n")
        file.close()
    
    def send_message(self, mail_app, destination, subject, body):
        mail = mail_app.CreateItem(0)
        mail.To = destination
        mail.Subject = subject
        mail.Body = body
        mail.Send()
        
    def send_emails(self):
        for name in self.reviews:
            self.send_message(self.outlook, self.reviews[name].contact, "FSE100 Peer Review Summary", self.format_review(self.reviews[name]))
            
        
#reader = Eval_Reader("FSE100 Assembly Line Project Peer Evaluation II (Responses).xlsx")
reader = Eval_Reader("FSE100A F18.xlsx")
reader.update_contacts()
reader.update_reviews()
reader.finalize_reviews()
reader.write_email_file("test_out.txt")
reader.send_emails()

